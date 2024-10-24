# ----------------
# App Improvements
# ----------------

# 1 Separate button for daily deposits to store list
# 5 when you make add_deposits_to_store remember to check the previous thursday's store books
# 5 add_ins_outs needs to account for people phasing up or down on tuesdays
#7 generate_deposits_sheet can be updated to make the totals cells dynamically update as deposits are added
#8 add_withdrawals should check that the patient has the money first
#9 Make a NO REFILL check for loading money onto account

# -------
# Imports
# -------

import os
import sys
import openpyxl
import pyodbc
from PyQt5.QtWidgets import (QApplication, QMainWindow,
                             QPushButton, QVBoxLayout,
                             QHBoxLayout, QTextEdit,
                             QLabel, QWidget, QGroupBox)
from PyQt5.QtCore import Qt
import pandas as pd
import xlwings as xw
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import json
import requests
import time
from API import API_KEY, API_PIN, API_PASSWORD, API_URL

# --------------
# ALL FILE PATHS
# --------------

# Function to load the configuration file
def load_config(config_file):
    with open(config_file, 'r') as file:
        return json.load(file)

config = load_config('config_work.json') # Used at work
#config = load_config('config_home.json') # Used at home

# Retrieve the base directories from the configuration file
excel_base_dir = config.get('excel_base_dir')
access_base_dir = config.get('access_base_dir')

# Construct the file paths dynamically using the base directories
database_path = os.path.join(access_base_dir, 'Client Trust.accdb')
auto_deposits_path = os.path.join(excel_base_dir, '2024 Deposits', 'Automated-Deposits-Sheet.xlsx')
auto_withdrawals_path = os.path.join(excel_base_dir, '2024 Withdrawals', 'Automated-Withdrawals-Sheet.xlsx')
auto_ins_outs_path = os.path.join(excel_base_dir, 'Ins N Outs', 'Automated-InsOuts.xlsx')
store_list_folder_path = os.path.join(excel_base_dir, 'Store List 2024')
linked_to_access_path = os.path.join(excel_base_dir, 'Store List 2024', 'Store List Linked To Access.xlsx')
deposits_folder_path = os.path.join(excel_base_dir, '2024 Deposits')
withdrawals_folder_path = os.path.join(excel_base_dir, '2024 Withdrawals')
new_store_folder_path = os.path.join(excel_base_dir,'New Store List')
quarters_folder_path = os.path.join(excel_base_dir, 'Quarters')

# Database Connection String
connection_string = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + database_path + ';'
)

# -----------------
# Main Window Class
# -----------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Set up the main window
        self.setWindowTitle("Client Trust Management")
        self.setGeometry(200, 200, 600, 400)

        # Main layout
        main_layout = QVBoxLayout()

        # Horizontal layout to hold the groups
        button_layout = QHBoxLayout()

        # 1. Organization and Deposits Group
        org_deposits_group = QGroupBox("Organization and Deposits")
        org_deposits_group.setAlignment(Qt.AlignCenter)
        org_deposits_layout = QVBoxLayout()

        self.ins_and_outs_button = QPushButton("Add Ins && Outs to Access")
        self.ins_and_outs_button.clicked.connect(self.add_ins_outs)
        org_deposits_layout.addWidget(self.ins_and_outs_button)

        self.generate_deposit_sheet_button = QPushButton("Generate New Deposits Sheet")
        self.generate_deposit_sheet_button.clicked.connect(self.generate_deposits_sheet)
        org_deposits_layout.addWidget(self.generate_deposit_sheet_button)

        org_deposits_group.setLayout(org_deposits_layout)
        button_layout.addWidget(org_deposits_group)

        # 2. Store List Group
        store_list_group = QGroupBox("Store List")
        store_list_group.setAlignment(Qt.AlignCenter)
        store_list_layout = QVBoxLayout()

        self.add_new_patients_button = QPushButton("Add New/Dephased Patients to Comcash")
        self.add_new_patients_button.clicked.connect(self.new_patients_to_comcash)
        store_list_layout.addWidget(self.add_new_patients_button)

        self.remove_patients_comcash_button = QPushButton("Delete Discharged/2nd Phase Patients from Comcash")
        self.remove_patients_comcash_button.clicked.connect(self.delete_patients_from_comcash)
        store_list_layout.addWidget(self.remove_patients_comcash_button)

        self.store_list_button = QPushButton("Generate Today's Store List")
        self.store_list_button.clicked.connect(self.generate_store_list)
        store_list_layout.addWidget(self.store_list_button)

        self.replenish_books_thurs_button = QPushButton("Store Balances to $100")
        self.replenish_books_thurs_button.clicked.connect(self.replenish_store_balances_thurs)
        store_list_layout.addWidget(self.replenish_books_thurs_button)

        self.new_store_list_button = QPushButton("Generate New Store List")
        self.new_store_list_button.clicked.connect(self.generate_new_store_list)
        store_list_layout.addWidget(self.new_store_list_button)

        self.add_deposits_to_store_button = QPushButton("Add Daily Deposits to New Store List")
        self.add_deposits_to_store_button.clicked.connect(self.add_daily_deposits_to_store_list)
        store_list_layout.addWidget(self.add_deposits_to_store_button)

        self.replenish_balances_store_button = QPushButton("Replenish Store Balances to $100")
        self.replenish_balances_store_button.clicked.connect(self.replenish_new_store_balances)
        store_list_layout.addWidget(self.replenish_balances_store_button)

        store_list_group.setLayout(store_list_layout)
        button_layout.addWidget(store_list_group)

        # 3. Balancing the Client Trust Database Group
        balance_group = QGroupBox("Balancing Client Trust")
        balance_group.setAlignment(Qt.AlignCenter)
        balance_layout = QVBoxLayout()

        self.generate_withdrawal_sheet_button = QPushButton("Generate New Withdrawals Sheet")
        self.generate_withdrawal_sheet_button.clicked.connect(self.generate_withdrawals_sheet)
        balance_layout.addWidget(self.generate_withdrawal_sheet_button)

        self.deposit_button = QPushButton("Add Deposits to Access")
        self.deposit_button.clicked.connect(self.add_deposits)
        balance_layout.addWidget(self.deposit_button)

        self.withdrawal_button = QPushButton("Add Withdrawals to Access")
        self.withdrawal_button.clicked.connect(self.add_withdrawals)
        balance_layout.addWidget(self.withdrawal_button)

        self.discharge_button = QPushButton("Discharge $0.00 Balances")
        self.discharge_button.clicked.connect(self.discharge_patients)
        balance_layout.addWidget(self.discharge_button)

        balance_group.setLayout(balance_layout)
        button_layout.addWidget(balance_group)

        # Add the horizontal layout to the main layout
        main_layout.addLayout(button_layout)

        # Text box for displaying results
        self.result_box = QTextEdit()
        self.result_box.setReadOnly(True)
        main_layout.addWidget(self.result_box)

        # Container widget
        container = QWidget()
        container.setLayout(main_layout)

        # Set the central widget to the container
        self.setCentralWidget(container)

    def discharge_patients(self):
        # Default connection
        connection = None
        try:
            # Establish the connection
            connection = pyodbc.connect(connection_string)

            # Create a cursor object using the connection
            cursor = connection.cursor()
            query = '''
                SELECT FirstName, LastName, Phase, [Sum Of DepositAmount], [Sum Of WithdrawalAmount]
                FROM Balance
                '''
            cursor.execute(query)
            rows = cursor.fetchall()

            # Clear previous results
            self.result_box.clear()

            # Keep count of discharged patients
            patient_discharge_count = 0

            # Iterate over the rows and print them
            for row in rows:
                # Get name
                first_name = row[0]
                last_name = row[1]

                # Get phase
                phase = int(row[2])

                # Get deposits and withdrawals
                deposit = float(row[3])
                withdrawal = float(row[4])

                # Calculate balance
                balance = deposit - withdrawal

                # Mark Discharged if patient is phase 4 and balance = 0.00
                if phase == 4 and balance == 0.0000:
                    self.result_box.append(f"Discharging: {last_name}, {first_name}, "
                                           f"Phase: {phase}, Balance: {balance}")
                    update_query = '''
                        UPDATE Clients
                        SET Discharged = True
                        WHERE FirstName = ? AND LastName = ?
                    '''
                    cursor.execute(update_query, (first_name, last_name))
                    connection.commit()
                    patient_discharge_count = patient_discharge_count + 1

            # If no patients were discharge, print message for the user
            if patient_discharge_count == 0:
                self.result_box.append(f"No patients met the discharge criteria.")

        # Print error if found
        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def add_deposits(self):
        # Path to Excel file
        excel_path = auto_deposits_path

        # Load the Excel file into a pandas DataFrame
        df = pd.read_excel(excel_path)
        connection = None

        try:
            # Establish the connection to Access Database
            connection = pyodbc.connect(connection_string)

            # Create a cursor object using the Access connection
            cursor = connection.cursor()

            # Clear previous results
            self.result_box.clear()

            # Iterate over each row in the Excel Dataframe
            for index, row in df.iterrows():
                excel_first_name = row['FirstName']
                excel_last_name = row['LastName']
                excel_type = row['Type']
                excel_amount = row['Amount']

                # Query the Access database for the client
                query = """
                                SELECT ClientID, FirstName, LastName, Phase
                                FROM Clients
                                WHERE FirstName = ? AND LastName = ?
                            """

                cursor.execute(query, (excel_first_name, excel_last_name))
                client = cursor.fetchone()

                # Check if the client exists
                if not client:
                    # If no client is found, print a message
                    self.result_box.append(
                        f"{excel_first_name} {excel_last_name} not found in Access. Deposit was not added.")
                    self.result_box.append("")
                    continue

                client_id, first_name, last_name, phase = client

                # Check if the client's phase is 4
                if phase == 4:
                    # If phase is 4, print a message and skip the deposit
                    self.result_box.append(
                        f"{excel_first_name} {excel_last_name} found in Access, but Phase: 4. Deposit was not added.")
                    self.result_box.append("")
                    continue

                # Query the Access database
                query = """
                    SELECT ClientID, FirstName, LastName
                    FROM Clients
                    WHERE FirstName = ? AND LastName = ?
                """

                cursor.execute(query, (excel_first_name, excel_last_name))
                client = cursor.fetchone()

                if client:
                    client_id = client[0]

                    # Get today's date for the transaction
                    today_date = datetime.today().strftime('%m/%d/%Y')

                    # Check if a matching transaction already exists
                    check_query = """
                    SELECT TransactionID
                    FROM Transactions
                    WHERE TransactionDate = ? AND TransactionDescription = ? 
                    AND DepositAmount = ? AND ClientID = ?
                    """
                    cursor.execute(check_query, (today_date, excel_type, excel_amount, client_id))
                    existing_transaction = cursor.fetchone()

                    if existing_transaction:
                        # If transaction already exists, print message
                        self.result_box.append(
                            f"Transaction already exists for {excel_first_name} {excel_last_name} "
                            f"with the amount {excel_amount} on {today_date}.")
                        self.result_box.append("")
                    else:
                        # Get the highest current transaction number from the Transactions table
                        max_transaction_query = "SELECT MAX(TransactionID) FROM Transactions"
                        cursor.execute(max_transaction_query)
                        max_transaction = cursor.fetchone()[0]  # Fetch the current maximum

                        # Create the new transaction number by adding 1
                        new_transaction_number = max_transaction + 1

                        # Get today's date for the transaction
                        today_date = datetime.today().strftime('%m/%d/%Y')

                        # Insert the new transaction entry
                        insert_query = """
                                        INSERT INTO Transactions (TransactionID, TransactionDate, 
                                        TransactionDescription, DepositAmount, WithdrawalAmount, ClientID)
                                        VALUES (?, ?, ?, ?, ?, ?)
                                    """

                        deposit_amount = excel_amount
                        withdrawal_amount = 0.00

                        cursor.execute(insert_query, (new_transaction_number, today_date,
                                                      excel_type, deposit_amount, withdrawal_amount, client_id))

                        self.result_box.append(
                            f"Deposited: {deposit_amount} to {excel_first_name} "
                            f"{excel_last_name}'s account on {today_date}.")
                        self.result_box.append("")

                        # Commit the transaction
                        connection.commit()

        except FileNotFoundError:
            self.result_box.setText("Deposits Excel file not found.")

        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")

        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

        finally:
            if connection:
                connection.close()

    def add_withdrawals(self):
        # Path to Excel file
        excel_path = auto_withdrawals_path
        connection = None
        try:
            # Load the Excel file into a pandas DataFrame
            df = pd.read_excel(excel_path)

            # Establish the connection to Access Database
            connection = pyodbc.connect(connection_string)

            # Create a cursor object using the Access connection
            cursor = connection.cursor()

            # Clear previous results
            self.result_box.clear()

            # Iterate over each row in the Excel Dataframe
            for index, row in df.iterrows():
                excel_first_name = row['FirstName']
                excel_last_name = row['LastName']
                excel_type = row['Type']
                excel_amount = row['Amount']

                # Query the Access database for the client
                query = """
                                SELECT ClientID, FirstName, LastName, Phase
                                FROM Clients
                                WHERE FirstName = ? AND LastName = ?
                            """

                cursor.execute(query, (excel_first_name, excel_last_name))
                client = cursor.fetchone()

                # Check if the client exists
                if not client:
                    # If no client is found, print a message
                    self.result_box.append(
                        f"{excel_first_name} {excel_last_name} not found in Access. Withdrawal was not added.")
                    self.result_box.append("")
                    continue

                client_id, first_name, last_name, phase = client

                # Check if the client's phase is 4
                if phase == 4:
                    # If phase is 4, print a message and skip the deposit
                    self.result_box.append(
                        f"{excel_first_name} {excel_last_name} found in Access, but Phase: 4. Withdrawal was not added.")
                    self.result_box.append("")
                    continue

                # Get today's date for the transaction
                today_date = datetime.today().strftime('%m/%d/%Y')

                # Check if a matching transaction already exists
                check_query = """
                            SELECT TransactionID
                            FROM Transactions
                            WHERE TransactionDate = ? AND TransactionDescription = ? 
                            AND WithdrawalAmount = ? AND ClientID = ?
                            """
                cursor.execute(check_query, (today_date, excel_type, excel_amount, client_id))
                existing_transaction = cursor.fetchone()

                if existing_transaction:
                    # If transaction already exists, print message
                    self.result_box.append(
                        f"Transaction already exists for {excel_first_name} {excel_last_name} "
                        f"with the amount {excel_amount} on {today_date}.")
                    self.result_box.append("")
                else:
                    # Get the highest current transaction number from the Transactions table
                    max_transaction_query = "SELECT MAX(TransactionID) FROM Transactions"
                    cursor.execute(max_transaction_query)
                    max_transaction = cursor.fetchone()[0]  # Fetch the current maximum

                    # Create the new transaction number by adding 1
                    new_transaction_number = max_transaction + 1

                    # Insert the new transaction entry
                    insert_query = """
                                                INSERT INTO Transactions (TransactionID, TransactionDate, 
                                                TransactionDescription, DepositAmount, WithdrawalAmount, ClientID)
                                                VALUES (?, ?, ?, ?, ?, ?)
                                            """

                    deposit_amount = 0.00
                    withdrawal_amount = excel_amount

                    cursor.execute(insert_query, (new_transaction_number, today_date,
                                                  excel_type, deposit_amount, withdrawal_amount, client_id))

                    self.result_box.append(
                        f"Withdrew: {withdrawal_amount} from {excel_first_name} "
                        f"{excel_last_name}'s account on {today_date}.")
                    self.result_box.append("")

                    # Commit the transaction
                    connection.commit()

        except FileNotFoundError:
            self.result_box.setText("Withdrawal Excel file not found.")

        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")

        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def add_ins_outs(self):
        # Default connection
        connection = None

        try:
            # Ins and Outs file path
            file_path = auto_ins_outs_path

            # Read into a dataframe
            df = pd.read_excel(file_path)

            # Establish connection to Access Database
            connection = pyodbc.connect(connection_string)
            cursor = connection.cursor()

            # Step 1: Get the max ClientID from the Clients table
            max_id_query = "SELECT MAX(ClientID) FROM Clients"
            cursor.execute(max_id_query)
            max_client_id = cursor.fetchone()[0]

            self.result_box.clear()

            # Step 2: Insert new clients from admissions_df into the Clients table
            for index, row in df.iterrows():
                entry_type = row['Type']
                first_name = row['FirstName']
                last_name = row['LastName']

                if entry_type == 'A':
                    # Check if the client already exists in the database
                    client_check_query = """
                            SELECT ClientID FROM Clients
                            WHERE FirstName = ? AND LastName = ?
                        """

                    cursor.execute(client_check_query, first_name, last_name)
                    existing_client = cursor.fetchone()

                    if existing_client is not None:
                        self.result_box.append(f"{first_name} {last_name} is already in Access Database.")
                        self.result_box.append(
                            "Check with records and manually enter the patient into Access Database.")
                        self.result_box.append("")
                    else:
                        client_id = max_client_id + 1  # Incrementing for each new client
                        max_client_id = client_id
                        first_name = row['FirstName']
                        last_name = row['LastName']
                        phase = 1
                        discharged = False
                        comments = None  # No comments as per your request
                        contract = row['Contract']

                        # SQL query to insert the new client
                        insert_query = """
                            INSERT INTO Clients (ClientID, FirstName, LastName, Phase, Discharged, Comments, Contract)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                            """

                        # Execute the query for each client
                        cursor.execute(insert_query, client_id, first_name, last_name, phase, discharged, comments,
                                       contract)

                        # Get the max Transaction ID from the Transactions table
                        max_transaction_id_query = "SELECT MAX(TransactionID) FROM Transactions"
                        cursor.execute(max_transaction_id_query)
                        max_transaction_id = cursor.fetchone()[0]

                        if max_transaction_id is None:
                            max_transaction_id = 1
                        else:
                            max_transaction_id += 1

                        # Transaction details
                        transaction_id = max_transaction_id
                        transaction_date = datetime.now().strftime('%m/%d/%Y')  # Today's date
                        transaction_description = 'Beginning Balance'
                        deposit_amount = 0.00
                        withdrawal_amount = 0.00

                        # SQL query to insert a new transaction
                        transaction_query = """
                                INSERT INTO Transactions (TransactionID, TransactionDate, TransactionDescription, DepositAmount, WithdrawalAmount, ClientID)
                                VALUES (?, ?, ?, ?, ?, ?)
                                """

                        # Execute the query to add the new transaction
                        cursor.execute(transaction_query, transaction_id, transaction_date, transaction_description,
                                       deposit_amount,
                                       withdrawal_amount, client_id)

                        self.result_box.append(f"{first_name} {last_name} added to Access Database.")
                        self.result_box.append("")
                elif entry_type == 'D':
                    # Step 2: Insert Discharges from admissions_df into the Clients table on Access Database
                    reason_for_discharge = row['ReasonForDischarge']

                    # Find the Access ClientID by matching first and last names
                    client_id_query = """
                        SELECT ClientID, Phase FROM Clients
                        WHERE FirstName = ? AND LastName = ?
                    """

                    cursor.execute(client_id_query, first_name, last_name)
                    client_data = cursor.fetchone()

                    if client_data is not None:
                        client_id, current_phase = client_data

                        # Check if the patient is already phase 4
                        if int(current_phase) == 4:
                            self.result_box.append(f"{first_name} {last_name} has already been discharged.")
                            self.result_box.append("")
                        else:
                            # Get the max TransactionID from the Transactions table
                            max_transaction_id_query = """
                                    SELECT MAX(TransactionID) FROM Transactions
                                    """
                            cursor.execute(max_transaction_id_query)
                            max_transaction_id = cursor.fetchone()[0]

                            max_transaction_id = max_transaction_id + 1

                            # Transaction details for discharge
                            transaction_id = max_transaction_id
                            transaction_date = datetime.now().strftime("%m/%d/%Y")
                            transaction_description = reason_for_discharge
                            deposit_amount = 0.00
                            withdrawal_amount = 0.00

                            # SQL query to insert a new discharge transaction
                            transaction_query = """
                                        INSERT INTO Transactions (TransactionID, TransactionDate, TransactionDescription, 
                                        DepositAmount, WithdrawalAmount, ClientID)
                                        VALUES (?, ?, ?, ?, ?, ?)
                                    """

                            cursor.execute(transaction_query, transaction_id, transaction_date,
                                           transaction_description, deposit_amount, withdrawal_amount, client_id)

                            update_phase_query = """
                                        UPDATE Clients
                                        SET Phase = 4
                                        WHERE ClientID = ?
                                    """
                            cursor.execute(update_phase_query, client_id)

                            self.result_box.append(f"{first_name} {last_name} discharged from Access Database.")
                            self.result_box.append("")
                    else:
                        self.result_box.append(
                            f"Client {first_name} {last_name} not found in Access Database. No discharge entry added.")
                        self.result_box.append("")
                else:
                    self.result_box.append(f"Invalid entry type for {first_name} {last_name}")

            # Commit the transaction
            connection.commit()

        except FileNotFoundError:
            self.result_box.setText("Ins n Outs Excel file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def replenish_store_balances_thurs(self):
        # Default connection
        connection = None

        try:
            connection = pyodbc.connect(connection_string)
            cursor = connection.cursor()

            # Get today's date in the format MM-DD-YY
            today = datetime.now().strftime("%m-%d-%y")

            # Construct the expected file name
            expected_file_name = f"Store List_{today}.xlsx"

            # Build the full file path
            folder_path = store_list_folder_path
            excel_file_path = os.path.join(folder_path, expected_file_name)

            self.result_box.clear()

            try:
                # Use xlwings to open the workbook and recalculate all formulas
                app = xw.App(visible=False)  # Run Excel in the background
                workbook = xw.Book(excel_file_path)
                worksheet = workbook.sheets[0]

                # Recalculate the workbook to ensure all formulas are updated
                workbook.api.RefreshAll()  # Refresh all data connections and formulas
                worksheet.api.Calculate()  # Recalculate worksheet formulas

                # Iterate over the rows, starting at row 2
                for row in range(2, worksheet.range('A1').current_region.last_cell.row + 1):
                    last_name = worksheet.range(f'A{row}').value
                    first_name = worksheet.range(f'B{row}').value
                    store_balance = worksheet.range(f'G{row}').value  # Get the value from formula in Column G

                    # Stop the loop if last_name is None
                    if last_name is None:
                        break
                    # Query the Access database for the matching LastName and FirstName
                    query = """
                    SELECT [Sum of DepositAmount], [Sum of WithdrawalAmount]
                    FROM Balance
                    WHERE [LastName] = ? AND [FirstName] = ?
                    """

                    cursor.execute(query, (last_name, first_name))
                    result = cursor.fetchone()

                    if result:
                        deposit_sum = result[0] if result[0] is not None else 0
                        withdrawal_sum = result[1] if result[1] is not None else 0

                        balance = float(deposit_sum - withdrawal_sum)
                        current_add_col = worksheet.range(f'F{row}').value \
                            if worksheet.range(f'F{row}').value is not None else 0

                        store_balance = float(store_balance)

                        # Only proceed if balance > 0 to avoid negative values being added
                        if balance > 0:
                            # Check if adding balance exceeds 100
                            if (balance + store_balance) > 100:
                                add_amount = 100 - store_balance - current_add_col
                                worksheet.range(f'F{row}').value = current_add_col + add_amount
                                self.result_box.append(
                                    f"Added {add_amount} to {first_name} {last_name}'s store balance.")
                            else:
                                worksheet.range(f'F{row}').value = current_add_col + balance
                                self.result_box.append(f"Added {balance} to {first_name} {last_name}'s store balance.")

                # Save the workbook to store recalculated values (optional)
                workbook.save()
                workbook.close()
                app.quit()

            except Exception as e:
                self.result_box.setText(f"Error loading Excel file: {e}")
                return

        except FileNotFoundError:
            self.result_box.setText("Store List Excel file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def generate_store_list(self):
        # Default connection
        connection = None
        try:
            # Create a connection to the Access database
            connection = pyodbc.connect(connection_string)
            cursor = connection.cursor()

            # Query to fetch clients with Phase = 1
            cursor.execute("SELECT LastName, FirstName FROM Clients WHERE Phase = '1'"
                           "ORDER BY LastName ASC, FirstName ASC")

            # Fetch all results
            clients = cursor.fetchall()

            # Set the folder path to save the new Store List workbook
            folder_path = store_list_folder_path

            # Get today's data and format it as MM-DD-YY
            today = datetime.today().strftime('%m-%d-%y')

            # Create the file name with today's date
            file_name = f'Store List_{today}.xlsx'

            # Full path to save the workbook
            file_path = os.path.join(folder_path, file_name)

            # Create a new blank Store List and select the active sheet
            sl = openpyxl.Workbook()
            ws = sl.active

            # Add headers
            ws['A1'] = 'LastName'
            ws['B1'] = 'FirstName'
            ws['C1'] = 'Store List Balance'
            ws['D1'] = 'Spent At Store'
            ws['E1'] = 'Quarters'
            ws['F1'] = 'Add to List'
            ws['G1'] = 'Balance'

            # Load the linked Excel file (Store List_Linked_To_Access)
            linked_file_path = linked_to_access_path
            linked_wb = openpyxl.load_workbook(linked_file_path, data_only=True)
            linked_ws = linked_wb.active

            # Iterate over clients and populate new file
            for idx, client in enumerate(clients, start=2):
                last_name = client[0]
                first_name = client[1]

                # Add names to the new sheet
                ws[f'A{idx}'] = last_name  # LastName in column A
                ws[f'B{idx}'] = first_name  # FirstName in column B

                # Search for the matching name in the linked file
                for row in linked_ws.iter_rows(min_row=2, values_only=True):  # Start from row 2, skip header
                    linked_last_name = row[0]
                    linked_first_name = row[1]
                    linked_balance = row[6]  # Column G (balance) is the 7th column, index 6

                    # Check if the names match
                    if last_name.lower() == linked_last_name.lower() and first_name.lower() == linked_first_name.lower():
                        ws[f'C{idx}'] = linked_balance  # Add the balance to the new file (column C)
                        break  # Stop searching once the match is found

                # Add formula to the G column (Balance) for the current row
                ws[f'G{idx}'] = f'=C{idx}-D{idx}-E{idx}+F{idx}'

            # After all clients are added, find the row after the last client
            last_row = len(clients) + 2  # Row after the last client
            first_client_row = 2  # First client always starts at row 2

            # Add the SUM formulas from the first to the last client
            ws[f'C{last_row}'] = f"=SUM(C{first_client_row}:C{last_row - 1})"
            ws[f'D{last_row}'] = f"=SUM(D{first_client_row}:D{last_row - 1})"
            ws[f'E{last_row}'] = f"=SUM(E{first_client_row}:E{last_row - 1})"
            ws[f'F{last_row}'] = f"=SUM(F{first_client_row}:F{last_row - 1})"
            ws[f'G{last_row}'] = f"=SUM(G{first_client_row}:G{last_row - 1})"

            # Create the Left list of clients
            left_title_row = len(clients) + 4

            # Merge columns A and B in the left_title_row
            ws.merge_cells(f'A{left_title_row}:B{left_title_row}')

            # Add the text "Left" in the merged cell
            ws[f'A{left_title_row}'] = "Left"
            ws[f'C{left_title_row}'] = "Store List Balance"
            ws[f'D{left_title_row}'] = "Spent At Store"
            ws[f'E{left_title_row}'] = "Quarters"
            ws[f'F{left_title_row}'] = "Balance"

            # Apply yellow background to the merged cell
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws[f'A{left_title_row}'].fill = yellow_fill

            # Center the text in the merged cell
            center_alignment = Alignment(horizontal="center", vertical="center")
            ws[f'A{left_title_row}'].alignment = center_alignment

            # Create a set of full names (LastName, FirstName) for easy comparison
            client_names = {(client[0].lower(), client[1].lower()) for client in clients}

            # Start filling the "Left" list, one row below the "Left" title row
            left_list_row = left_title_row + 1

            # Iterate through the linked Excel file and find names not in the clients list
            for row in linked_ws.iter_rows(min_row=2, values_only=True):
                linked_last_name = row[0]
                linked_first_name = row[1]
                linked_balance = row[6]

                # Only proceed if both last name and first name are not None
                if linked_last_name is not None and linked_first_name is not None:
                    # Compare full name from linked list with client names
                    if (linked_last_name.lower(), linked_first_name.lower()) not in client_names:
                        # Add unmatched name to the "Left" list
                        ws[f'A{left_list_row}'] = linked_last_name
                        ws[f'B{left_list_row}'] = linked_first_name
                        # Populate column C with the balance (Column G in linked file)
                        ws[f'C{left_list_row}'] = linked_balance
                        ws[f'F{left_list_row}'] = f'=C{left_list_row}-D{left_list_row}-E{left_list_row}'
                        left_list_row += 1  # Move to the next row for the next unmatched name

            # After all clients are added, find the row after the last client
            last_left_row = left_list_row
            first_left_client_row = left_title_row + 1

            # After all clients are added, check if any clients were added to the "Left" list
            if left_list_row > left_title_row + 1:
                # Add the SUM formulas from the first to the last client
                ws[f'C{last_left_row}'] = f"=SUM(C{first_left_client_row}:C{last_left_row - 1})"
                ws[f'D{last_left_row}'] = f"=SUM(D{first_left_client_row}:D{last_left_row - 1})"
                ws[f'E{last_left_row}'] = f"=SUM(E{first_left_client_row}:E{last_left_row - 1})"
                ws[f'F{last_left_row}'] = f"=SUM(F{first_left_client_row}:F{last_left_row - 1})"
            else:
                # If no clients were added, set the cells to 0.00
                ws[f'C{left_list_row}'] = 0.00
                ws[f'D{left_list_row}'] = 0.00
                ws[f'E{left_list_row}'] = 0.00
                ws[f'F{left_list_row}'] = 0.00

            # Title the final Calculation row
            final_calc_row = last_left_row + 2

            ws[f'C{final_calc_row}'] = "Total"
            ws[f'D{final_calc_row}'] = "Spent"
            ws[f'E{final_calc_row}'] = "Quarters"
            ws[f'F{final_calc_row}'] = "Left"
            ws[f'G{final_calc_row}'] = "Added"
            ws[f'H{final_calc_row}'] = "Balance"

            # Highlight the final calculation row
            ws[f'C{final_calc_row}'].fill = yellow_fill
            ws[f'D{final_calc_row}'].fill = yellow_fill
            ws[f'E{final_calc_row}'].fill = yellow_fill
            ws[f'F{final_calc_row}'].fill = yellow_fill
            ws[f'G{final_calc_row}'].fill = yellow_fill
            ws[f'H{final_calc_row}'].fill = yellow_fill

            # Enter the formulas
            ws[f'C{final_calc_row + 1}'] = f'=C{last_row}+C{last_left_row}'
            ws[f'D{final_calc_row + 1}'] = f'=D{last_row}+D{last_left_row}'
            ws[f'E{final_calc_row + 1}'] = f'=E{last_row}+E{last_left_row}'
            ws[f'F{final_calc_row + 1}'] = f'=F{last_left_row}'
            ws[f'G{final_calc_row + 1}'] = f'=F{last_row}'
            ws[f'H{final_calc_row + 1}'] = (f'=C{final_calc_row + 1}-D{final_calc_row + 1}'
                                            f'-E{final_calc_row + 1}-F{final_calc_row + 1}+G{final_calc_row + 1}')

            # Save and open the new workbook
            sl.save(file_path)
            os.startfile(file_path)

            # Clear previous results and print confirmation
            self.result_box.clear()
            self.result_box.append(f"{file_name} created.")

        except FileNotFoundError:
            self.result_box.setText("Linked file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def generate_deposits_sheet(self):
        try:
            # Path to the folder containing the files
            folder_path = deposits_folder_path

            # Get the list of files in the folder
            files = os.listdir(folder_path)

            # Filter only the files with the correct prefix
            deposit_files = [f for f in files if f.startswith("Deposits for Client Trust")]

            if not deposit_files:
                self.result_box.setText("No deposit files found.")
                return

            # Sort files by the date in the format MM-DD-YY (removing .xlsx extension)
            deposit_files.sort(key=lambda x: datetime.strptime(x.split()[-1].replace('.xlsx', ''),
                                                               '%m-%d-%y'), reverse=True)
            # Get the latest file
            latest_file = deposit_files[0]

            # Create the full path to the source file
            source_file = os.path.join(folder_path, latest_file)

            # Get today's date
            today_date = datetime.now().strftime('%m-%d-%y')

            # Create the full path to the destination file with today's date
            destination_file = os.path.join(folder_path, f"Deposits for Client Trust {today_date}.xlsx")

            # Copy the file to the destination
            shutil.copy(source_file, destination_file)

            # Now modify the new copy to clear specified cells
            # Load the workbook
            workbook = openpyxl.load_workbook(destination_file)

            # Iterate over both Sheet1 and Sheet2
            for sheet_name in ['Sheet 1', 'Sheet 2']:
                sheet = workbook[sheet_name]

                # Edit cell B5 to contain the current date
                sheet['B5'].value = f"Date: {today_date}"

                # Clear columns B, D, F, H in rows 8 through 33
                for row in range(8, 29):
                    for col in ['B', 'D', 'F', 'H']:
                        sheet[f'{col}{row}'].value = None  # Clear the content

            # Save the modified workbook
            workbook.save(destination_file)

            # Update the result box to show success
            self.result_box.setText(f"Successfully created file: {destination_file}")

            # Open the newly created file
            os.startfile(destination_file)

        except FileNotFoundError:
            self.result_box.setText("Deposit file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Database error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

    def generate_withdrawals_sheet(self):
        try:
            # Path to the folder containing the files
            folder_path = withdrawals_folder_path

            # Get the list of files in the folder
            files = os.listdir(folder_path)

            # Filter only the files with the correct prefix
            withdrawal_files = [f for f in files if f.startswith("Withdrawals for Client Trust")]

            if not withdrawal_files:
                self.result_box.setText("No withdrawal files found.")
                return

            # Sort files by the date in the format MM-DD-YY (removing .xlsx extension)
            withdrawal_files.sort(key=lambda x: datetime.strptime(x.split()[-1].replace('.xlsx', ''),
                                                                  '%m-%d-%y'), reverse=True)
            # Get the latest file
            latest_file = withdrawal_files[0]

            # Create the full path to the source file
            source_file = os.path.join(folder_path, latest_file)

            # Get today's date
            today_date = datetime.now().strftime('%m-%d-%y')

            # Create the full path to the destination file with today's date
            destination_file = os.path.join(folder_path, f"Withdrawals for Client Trust {today_date}.xlsx")

            # Copy the file to the destination
            shutil.copy(source_file, destination_file)

            # Now modify the new copy to clear specified cells
            # Load the workbook
            workbook = openpyxl.load_workbook(destination_file)

            # Iterate over both Sheet1 and Sheet2
            for sheet_name in ['Sheet 1', 'Sheet 2']:
                sheet = workbook[sheet_name]

                # Edit cell B5 to contain the current date
                sheet['B5'].value = f"Date: {today_date}"

                # Clear columns B, D, F, H in rows 8 through 33
                for row in range(8, 30):
                    for col in ['B', 'D', 'F', 'H']:
                        sheet[f'{col}{row}'].value = None  # Clear the content

            # Save the modified workbook
            workbook.save(destination_file)

            # Update the result box to show success
            self.result_box.setText(f"Successfully created file: {destination_file}")

            # Open the newly created file
            os.startfile(destination_file)

        except FileNotFoundError:
            self.result_box.setText("Withdrawal file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Database error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

    def new_patients_to_comcash(self):
        # Default connection
        connection = None

        try:
            # Establish the connection to Access Database
            connection = pyodbc.connect(connection_string)

            # Create a cursor object using the Access connection
            cursor = connection.cursor()

            # Query all phase 1 patients
            query = '''
                    SELECT FirstName, LastName, Phase
                    FROM Balance
                    WHERE Phase = '1';
            '''
            cursor.execute(query)

            rows = cursor.fetchall()

            # Establish connection to Comcash API
            api_client = APIClient()

            deleted_customer_list = api_client.get_customer_list(2, 4)
            active_customer_list = api_client.get_customer_list(1,4)

            # Clear previous results
            self.result_box.clear()

            customers_added = 0

            for row in rows:
                first_name = row.FirstName
                last_name = row.LastName

                # Check if the customer exists in the deleted_customer_list
                customer_in_deleted_list = False


                for deleted_customer in deleted_customer_list:
                    if deleted_customer['firstName'] == first_name and deleted_customer['lastName'] == last_name:
                        customer_in_deleted_list = True
                        break

                if not customer_in_deleted_list:
                    customer_in_active_list = False
                    for active_customer in active_customer_list:
                        if active_customer['firstName'] == first_name and active_customer['lastName'] == last_name:
                            customer_in_active_list = True
                            break

                    # Print appropriate messages based on customer presence
                    if not customer_in_active_list:
                        self.result_box.append(f"Creating a new customer: {first_name} {last_name}")
                        new_customer = api_client.create_new_customer(first_name, last_name)
                        api_client.update_customer_type(new_customer.get("id"))
                        customers_added = customers_added + 1
                else:
                    self.result_box.append(f"Customer {first_name} {last_name} is in the deleted customer list. "
                                           f"You must manually change their status to active.")

            if customers_added == 0:
                self.result_box.append("No patients were added to Comcash.")

        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def delete_patients_from_comcash(self):
        # Default connection
        connection = None

        try:
            # Establish the connection to Access Database
            connection = pyodbc.connect(connection_string)

            # Create a cursor object using the Access connection
            cursor = connection.cursor()

            # Query all phase 1 patients
            query = '''
                    SELECT FirstName, LastName, Phase
                    FROM Balance
                    WHERE Phase IN ('2', '3', '4');
            '''
            cursor.execute(query)

            rows = cursor.fetchall()

            # Establish connection to Comcash API
            api_client = APIClient()

            active_customer_list = api_client.get_customer_list(1,4)

            # Clear previous results
            self.result_box.clear()

            # For every row in the database query
            for row in rows:
                first_name = row.FirstName
                last_name = row.LastName

                # For every active customer in the list, if there is name match, delete the customer
                for active_customer in active_customer_list:
                    if active_customer['firstName'] == first_name and active_customer['lastName'] == last_name:
                        api_client.delete_customer(int(active_customer.get("id")))
                        self.result_box.append(f"{first_name} {last_name}'s account removed from Comcash.")

        except pyodbc.Error as e:
            self.result_box.setText(f"Error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")
        finally:
            if connection:
                connection.close()

    def generate_new_store_list(self):
        try:
            # Create a connection to the Comcash API
            api_client = APIClient()

            # Get the customer list
            customers = api_client.get_customer_list(1, 4)

            # Create a connection to the Access database
            connection = pyodbc.connect(connection_string)
            cursor = connection.cursor()

            # Query to fetch clients with Phase = 1
            cursor.execute("SELECT LastName, FirstName FROM Clients WHERE Phase = '1' "
                           "ORDER BY LastName ASC, FirstName ASC")

            # Fetch all results
            clients = cursor.fetchall()

            # Set the folder path to save the new Store List workbook
            store_folder_path = new_store_folder_path
            quarters_path = quarters_folder_path

            # Get today's date and format it as MM-DD-YY
            today = datetime.today().strftime('%m-%d-%y')

            # Get yesterday's date and format it as MM-DD-YY for the previous store list
            yesterday = (datetime.today() - timedelta(days=1)).strftime('%m-%d-%y')

            # Create the file names with today's date
            store_file_name = f'Store List_{today}.xlsx'
            previous_file_name = f'Store List_{yesterday}.xlsx'
            quarters_file_name = f'Quarters_{today}.xlsx'

            # Full path to save the workbook and previous file path
            store_file_path = os.path.join(store_folder_path, store_file_name)
            previous_file_path = os.path.join(store_folder_path, previous_file_name)
            quarters_file_path = os.path.join(quarters_path, quarters_file_name)

            # Check if the previous store list exists
            previous_data = {}

            # Open the workbook using xlwings
            with xw.App(visible=False) as app:
                wb = app.books.open(previous_file_path)
                ws = wb.sheets[0]  # Adjust if it's not the first sheet

                # Read the data starting from row 4
                for row_num in range(4, ws.api.UsedRange.Rows.Count + 1):
                    prev_last_name = ws.range(f'A{row_num}').value
                    prev_first_name = ws.range(f'B{row_num}').value
                    final_balance = ws.range(f'H{row_num}').value  # Assuming final balance is in column H

                    # Store in the dictionary
                    if prev_last_name and prev_first_name:
                        previous_data[(prev_last_name, prev_first_name)] = final_balance

                wb.close()

            if os.path.exists(store_file_path):
                print(f"File '{store_file_name}' already exists. No new file created.")
            else:
                # Create a new blank Store List and select the active sheet
                sl = openpyxl.Workbook()
                ws = sl.active

                # Create Title Header with today's date
                bold_side = Side(border_style="thick", color="000000")
                ws.merge_cells('A1:B1')
                ws['A1'] = "Store List"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                ws['A1'].border = Border(left=bold_side, top=bold_side, bottom=bold_side)
                ws['B1'].border = Border(top=bold_side, bottom=bold_side)
                ws['C1'] = f"{today}"
                ws['C1'].font = Font(bold=True, size=14)
                ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                ws['C1'].border = Border(right=bold_side, top=bold_side, bottom=bold_side)
                ws.column_dimensions['C'].width = 12

                # Create Table Headers
                ws['A3'] = "Last Name"
                ws['B3'] = "First Name"
                ws['C3'] = "Starting Balance"
                ws['D3'] = "Store Transactions"
                ws['E3'] = "Total Spent at Store"
                ws['F3'] = "Quarter Transactions"
                ws['G3'] = "Added to Store Balance"
                ws['H3'] = "Final Balance"

                row_num = 4  # Start at row 4
                for client in clients:
                    last_name = client.LastName
                    first_name = client.FirstName

                    # Insert Last Name in column A and First Name in column B
                    ws[f'A{row_num}'] = last_name
                    ws[f'B{row_num}'] = first_name

                    # Check if the client existed in the previous day's store list
                    final_balance = previous_data.get((last_name, first_name), 0)  # Default to 0 if no match found

                    # Set the final balance in the new list (Column C)
                    ws[f'C{row_num}'] = final_balance

                    # Get customer ID (you will need to adapt this to match your data model)
                    customer = next(
                        (c for c in customers if c.get('lastName') == last_name and c.get('firstName') == first_name),
                        None)

                    if customer:
                        customer_id = customer['id']

                        # Get sales for the customer
                        last_wednesday = datetime.now() - timedelta(days=6)
                        time_from = int(time.mktime(
                            datetime(last_wednesday.year, last_wednesday.month, last_wednesday.day, 0, 0,
                                     0).timetuple()))
                        time_to = int(time.mktime(
                            datetime(last_wednesday.year, last_wednesday.month, last_wednesday.day, 23, 59,
                                     59).timetuple()))

                        sales = api_client.get_customer_sales(customer_id, time_from, time_to)

                        # Initialize an empty list to store product details
                        product_list = []

                        # Initialize total payment
                        total_payment = 0.00

                        # Check if sales data exists
                        if not sales:
                            product_string = "No sales found"  # Set a default message if no sales are found
                        else:
                            # Process sales data (if any)
                            for sale in sales:
                                for product in sale['products']:
                                    product_title = product['title']
                                    product_price = product['price']
                                    # Add product details to the list
                                    product_list.append(f"{product_title} - ${product_price:.2f}")

                                # Directly access the payment dictionary
                                payment = sale.get('payment')
                                if isinstance(payment, dict) and 'totalPayedAmount' in payment:
                                    total_payment += float(payment['totalPayedAmount'])
                                else:
                                    self.result_box.append(
                                        f"Unexpected payment format or missing 'totalPayedAmount': {payment}")

                            # Join the list into a formatted string
                            product_string = '\n'.join(product_list)

                        # Set the product details string in column D and total sales in column E
                        ws[f'D{row_num}'] = product_string
                        ws[f'E{row_num}'] = total_payment

                    # Set the formula for Final Balance in column H
                    ws[f'H{row_num}'] = f"=C{row_num}-E{row_num}-F{row_num}+G{row_num}"

                    # Move to the next row
                    row_num += 1

                # Quarters sheet processing
                if os.path.exists(quarters_file_path):
                    quarters_wb = openpyxl.load_workbook(quarters_file_path)
                    quarters_ws = quarters_wb.active

                    # Iterate over the quarters sheet starting at row 2
                    for quarter_row in quarters_ws.iter_rows(min_row=2, max_row=quarters_ws.max_row, min_col=1,
                                                             max_col=3, values_only=True):
                        q_last_name, q_first_name, q_amount = quarter_row

                        # Track if the name was found in the store list
                        found_in_store_list = False

                        # Search for matching name in the store list (ws)
                        for store_row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=2,
                                                      values_only=False):
                            store_last_name_cell, store_first_name_cell = store_row
                            store_last_name = store_last_name_cell.value
                            store_first_name = store_first_name_cell.value

                            # If a match is found
                            if store_last_name == q_last_name and store_first_name == q_first_name:
                                found_in_store_list = True  # Mark as found

                                # Add the quarters amount (column C in quarters) to column F in store list
                                ws[f'F{store_last_name_cell.row}'] = q_amount
                                break

                        # If the name wasn't found, append to result box
                        if not found_in_store_list:
                            message = (f"{q_first_name} {q_last_name} has a ${q_amount} transaction on quarters sheet, "
                                       f"but the patient isn't on the store list")
                            self.result_box.append(message)

                # Determine the range of the table (from row 3 to the last row, columns A-H)
                table_range = f'A3:H{row_num - 1}'  # `row_num - 1` to account for the last row of data

                # Create a table and apply style
                table = Table(displayName="StoreListTable", ref=table_range)
                style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False, showLastColumn=False,
                                       showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style

                # Add the table to the worksheet
                ws.add_table(table)

                # Save the new workbook
                sl.save(store_file_path)
                print(f"New Store List '{store_file_name}' created successfully.")

        except FileNotFoundError:
            self.result_box.setText("Deposit file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Database error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

    def add_daily_deposits_to_store_list(self):
        try:
            # Read the Auto-Deposits Excel file using pandas
            df_deposits = pd.read_excel(auto_deposits_path)

            # Group deposits by 'FirstName' and 'LastName' and sum the 'Amount'
            df_grouped_deposits = df_deposits.groupby(['FirstName', 'LastName'], as_index=False)['Amount'].sum()

            # Determine today's date in the required format for the store list filename
            today = datetime.today().strftime('%m-%d-%y')
            store_list_file = os.path.join(new_store_folder_path, f'Store List_{today}.xlsx')

            # Open the store list excel file using xlwings
            app = xw.App(visible=False)  # Launch Excel in the background
            wb = app.books.open(store_list_file)
            sheet = wb.sheets[0]  # Assuming the relevant sheet is the first one

            # Function to calculate the previous Thursday
            def get_previous_thursday(start_date):
                offset = (start_date.weekday() - 3) % 7
                return start_date - timedelta(days=offset)

            # Get today's date and last Thursday's date
            today_date = datetime.today()
            last_thursday = get_previous_thursday(today_date)

            # Get yesterday's date
            yesterday = today_date - timedelta(days=1)

            # Generate the dates from last Thursday up to yesterday
            date_list = [last_thursday + timedelta(days=i) for i in range((yesterday - last_thursday).days + 1)]
            print(date_list)

            # Starting from row 4, get first and last name columns
            row = 4  # Starting row for names
            while True:
                store_last_name = sheet.range(f'A{row}').value
                store_first_name = sheet.range(f'B{row}').value

                # Stop if both first and last names are None
                if store_last_name is None and store_first_name is None:
                    break

                # Iterate through the grouped deposits and compare with store list names
                for index, deposit_row in df_grouped_deposits.iterrows():
                    first_name_deposit = deposit_row['FirstName']
                    last_name_deposit = deposit_row['LastName']
                    total_amount = deposit_row['Amount']

                    # Check if the name from the deposit file matches the store list names
                    if store_last_name == last_name_deposit and store_first_name == first_name_deposit:
                        # Calculate total_added_this_week by opening store lists for each day back to last Thursday
                        total_added_this_week = 0

                        # Loop through each date back to last Thursday, stopping at yesterday
                        for date in date_list:
                            date_str = date.strftime('%m-%d-%y')
                            store_file_path = os.path.join(new_store_folder_path, f'Store List_{date_str}.xlsx')

                            # Check if the file exists for this date
                            if os.path.exists(store_file_path):
                                try:
                                    # Open the file and access the sheet
                                    past_wb = app.books.open(store_file_path)
                                    past_sheet = past_wb.sheets[0]

                                    # Look for the person's 'Added' value in column G (assuming 'Added' is in column G)
                                    for r in range(4, past_sheet.range('A' + str(past_sheet.cells.last_cell.row)).end(
                                            'up').row + 1):
                                        past_last_name = past_sheet.range(f'A{r}').value
                                        past_first_name = past_sheet.range(f'B{r}').value

                                        # If the names match, sum up the "Added" column values
                                        if past_last_name == last_name_deposit and past_first_name == first_name_deposit:
                                            added_value = past_sheet.range(f'G{r}').value
                                            if added_value is not None:
                                                total_added_this_week += added_value
                                            break  # Exit the loop once a match is found

                                    # Close the past workbook
                                    past_wb.close()

                                except Exception as e:
                                    # Continue to the next file if there are issues opening this one
                                    print(f"Error accessing file {store_file_path}: {e}")
                                    continue

                        # Print the total_added_this_week along with the person's total deposits
                        print(
                            f"{last_name_deposit} {first_name_deposit}: {total_amount}, Total Added This Week: {total_added_this_week}")
                        break  # You can decide if you want to continue checking or break once matched

                row += 1  # Move to the next row in the store list

            # Clean up and close the Excel workbook
            wb.close()
            app.quit()

        except FileNotFoundError:
            self.result_box.setText("Deposit file or store list file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Database error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

    def replenish_new_store_balances(self):
        try:
            # Get today's date in MM-DD-YY format
            today = datetime.today().strftime('%m-%d-%y')

            # Set the expected file name for store list
            expected_file_name = f'Store List_{today}.xlsx'

            # Full path to the file (already set by new_store_folder_path)
            store_file_path = os.path.join(new_store_folder_path, expected_file_name)

            # Clear previous result box
            self.result_box.clear()

            # Check if the file exists
            if os.path.exists(store_file_path):
                try:
                    # Use xlwings to open the workbook
                    app = xw.App(visible=False)  # Run Excel in the background
                    workbook = xw.Book(store_file_path)
                    worksheet = workbook.sheets[0]

                    # Recalculate the workbook to ensure all formulas are updated
                    workbook.api.RefreshAll()  # Refresh all data connections and formulas
                    worksheet.api.Calculate()  # Recalculate worksheet formulas

                    # Establish connection to the Access database
                    connection = pyodbc.connect(connection_string)
                    cursor = connection.cursor()

                    # Start iterating from row 4 (assuming names always start at row 4)
                    row = 4
                    while True:
                        # Read the first name (column B) and last name (column A)
                        last_name = worksheet.range(f'A{row}').value
                        first_name = worksheet.range(f'B{row}').value

                        # Read add column and final balance column
                        current_final_balance = worksheet.range(f'H{row}').value
                        current_add_column = worksheet.range(f'G{row}').value
                        # Default to 0 if the "Added" column is None
                        if current_add_column is None:
                            current_add_column = 0

                        # If both first and last names are empty, stop the loop
                        if not first_name and not last_name:
                            break

                        # Query the Access database for the matching LastName and FirstName
                        query = """
                                    SELECT [Sum of DepositAmount], [Sum of WithdrawalAmount]
                                    FROM Balance
                                    WHERE [LastName] = ? AND [FirstName] = ?
                                    """
                        cursor.execute(query, (last_name, first_name))
                        result = cursor.fetchone()

                        if result:
                            sum_of_deposits, sum_of_withdrawals = result
                            if sum_of_deposits is None:
                                sum_of_deposits = 0  # Handle potential None values
                            if sum_of_withdrawals is None:
                                sum_of_withdrawals = 0  # Handle potential None values

                            # Calculate the access balance
                            access_balance = sum_of_deposits - sum_of_withdrawals
                            access_balance = float(access_balance)

                            amount_to_add = 0.00

                            if access_balance > 0.00:
                                if (current_add_column + current_final_balance + access_balance) <= 100:
                                    amount_to_add = (current_add_column + access_balance)
                                    worksheet.range(f'G{row}').value = amount_to_add
                                    self.result_box.append(f"Added ${amount_to_add} to {first_name} {last_name}'s account.")
                                else:
                                    amount_to_add = current_add_column + (100 - current_final_balance)
                                    if amount_to_add >= 0:
                                        worksheet.range(f'G{row}').value = amount_to_add
                                        self.result_box.append(f"Added ${amount_to_add} to {first_name} "
                                                               f"{last_name}'s account.")
                        else:
                            print(f"No data found for {first_name} {last_name} in the database.")

                        # Move to the next row
                        row += 1

                    # Close the workbook and app when done
                    workbook.save()
                    workbook.close()
                    app.quit()
                    connection.close()

                except Exception as e:
                    self.result_box.setText(f"Error loading Excel file: {e}")
                    return

            else:
                self.result_box.setText("You need to create today's store list before you can continue.")

        except FileNotFoundError:
            self.result_box.setText("Store file not found.")
        except pyodbc.Error as e:
            self.result_box.setText(f"Database error: {e}")
        except Exception as e:
            self.result_box.setText(f"Unexpected error: {e}")

# ----------------
# API Client CLass
# ----------------
class APIClient:
    def __init__(self, token_file="token_data.json"):
        self.api_key = API_KEY
        self.pin = API_PIN
        self.password = API_PASSWORD
        self.signin_url = f"{API_URL}/employee/auth/signin"
        self.customer_list_url = f"{API_URL}/employee/customer/list"
        self.create_customer_url = f"{API_URL}/employee/customer/create"
        self.update_customer_url = f"{API_URL}/employee/customer/update"
        self.update_balance_url = f"{API_URL}/employee/customer/updatePoints"
        self.delete_customer_url = f"{API_URL}/employee/customer/delete"
        self.get_sales_url = f"{API_URL}/employee/customer/sales"
        self.token_file = token_file
        self.token = None
        self.token_expiration = None
        self.load_token_from_file()

    def load_token_from_file(self):
        # Load token and expiration from a file if it exists
        if os.path.exists(self.token_file):
            with open(self.token_file, 'r') as file:
                data = json.load(file)
                self.token = data.get("token")
                self.token_expiration = data.get("token_expiration")
                print("Loaded token from file.")
        else:
            print("No token file found. Need to authenticate.")

    def save_token_to_file(self):
        # Save token and expiration to a file
        data = {
            "token": self.token,
            "token_expiration": self.token_expiration
        }
        with open(self.token_file, 'w') as file:
            json.dump(data, file)
        print("Token saved to file.")

    def authenticate(self):
        # Request a new bearer token
        payload = json.dumps({
            "openApiKey": self.api_key,
            "pin": self.pin,
            "password": self.password
        })
        headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(self.signin_url, headers=headers, data=payload)

        if response.status_code == 200:
            data = response.json()
            print(data)
            self.token = data.get("accessToken")
            expires_in = data.get("expiresIn")
            print(f"Expires in: {expires_in}")
            self.token_expiration = expires_in  # Track token expiration in seconds
            self.save_token_to_file()  # Save token to file after successful authentication
            print("Authenticated successfully. Token received.")
        else:
            print(f"Failed to authenticate. Status Code: {response.status_code}")

    def is_token_valid(self):
        # Check if token exists and hasn't expired
        if self.token and float(self.token_expiration) > time.time():
            return True
        return False

    def get_customer_list(self, status, customer_type):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "limit": 100000,
                "order": "asc",
                "type": customer_type,
                "status": status
            })

            response = requests.post(self.customer_list_url, headers=headers, data=payload)

            if response.status_code == 200:
                return response.json()
            else:
                print(f"Failed to fetch customer list. Status Code: {response.status_code}")
                return None

    def create_new_customer(self, first_name, last_name):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "countryPhoneCode": 1,
                "phone": 0000000000,
                "firstName": first_name,
                "lastName": last_name,
                "locationId": 1
            })

            response = requests.post(self.create_customer_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Customer created successfully.")
                return response.json()
            else:
                print(f"Failed to create customer. Status Code: {response.status_code}")
                return None

    def update_customer_type(self, customer_id):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "id": customer_id,
                "typeId": "4"
            })

            response = requests.post(self.update_customer_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Customer type updated successfully.")
                return response.json()
            else:
                print(f"Failed to update customer type. Status Code: {response.status_code}")
                return None

    def update_customer_balance(self, customer_id, balance):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "customerId": customer_id,
                "storeCredit": float(balance)
            })

            response = requests.post(self.update_balance_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Balance updated successfully.")
                return response.json()
            else:
                print(f"Failed to update balance. Status Code: {response.status_code}")
                return None

    def delete_customer(self, customer_id):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "customerId": customer_id
            })

            response = requests.post(self.delete_customer_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Customer successfully deleted.")
                return response.json()
            else:
                print(f"Failed to delete customer. Status Code: {response.status_code}")
                return None

    def update_customer_name(self, customer_id):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "id": customer_id,
                "firstName": "Testing",
                "lastName": "Again"
            })

            response = requests.post(self.update_customer_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Name updated successfully.")
                return response.json()
            else:
                print(f"Failed to update customer name. Status Code: {response.status_code}")
                return None

    def get_customer_sales(self, customer_id, time_from, time_to):
        if not self.is_token_valid():
            print("Token expired or invalid. Authenticating...")
            self.authenticate()

        if self.token:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {self.token}'
            }

            payload = json.dumps({
                "customerId": customer_id,
                "timeFrom": time_from,
                "timeTo": time_to
            })

            response = requests.post(self.get_sales_url, headers=headers, data=payload)

            if response.status_code == 200:
                print("Retrieved sales successfully.")
                return response.json()
            else:
                print(f"Failed to retrieve sales. Status Code: {response.status_code}")
                return None

# ------------------
# Main Program Logic
# ------------------

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())